from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    session,
    jsonify,
    make_response,
    get_flashed_messages,
)
import io
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import pandas as pd
import time
import json
import base64
import random
import requests
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import os
import uuid
import tempfile
import stat
from werkzeug.utils import secure_filename

from sku_matcher import judge_same_product
from pdf_parser import parse_pdf
from db import (
    init_db, save_parsed_document, list_invoices, delete_invoice,
    load_catalog_to_memory, get_catalog_df, refresh_catalog,
    get_user, list_users, add_user, set_user_active, set_user_role,
    increment_failed_attempts, reset_failed_attempts,
    log_login, get_login_history,
    save_template_db, get_template_db, list_templates_db,
    delete_template_db, rename_template_db, duplicate_template_db,
    get_template_versions_db, restore_template_version_db, count_templates_db,
    save_estimate_db, get_estimate_db, list_estimates_db,
    delete_estimate_db, duplicate_estimate_db,
    search_estimate_catalog, upsert_estimate_catalog,
    get_material_list_total,
)
import tempfile

# Additional imports for login functionality
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from itsdangerous import URLSafeTimedSerializer
from functools import wraps

# -------------------------------
# Application Configuration
# -------------------------------

import config
import data_utils as du

app = Flask(__name__)
init_db()
load_catalog_to_memory()
app.secret_key = config.SECRET_KEY


def _seed_estimate_catalog():
    """Seed the estimate catalog with template rows — runs only if catalog is empty."""
    try:
        from db import _turso_execute, _local_conn, USE_TURSO, upsert_estimate_catalog as _upsert
        check_sql = "SELECT COUNT(*) AS cnt FROM estimate_catalog"
        if USE_TURSO:
            rows = _turso_execute(check_sql)
            count = rows[0]["cnt"] if rows else 0
        else:
            with _local_conn() as conn:
                count = conn.execute(check_sql).fetchone()[0]
        if count > 0:
            return
        for section_name, _, rows in _TEMPLATE_SECTIONS:
            for row in rows:
                desc = row.get("description", "").strip()
                if not desc:
                    continue
                _upsert(
                    description   = desc,
                    unit_cost     = float(row.get("unit_cost") or 0),
                    comments      = row.get("comments", ""),
                    add_comments  = row.get("add_comments", ""),
                    category      = section_name,
                    estimate_name = "",
                )
    except Exception as e:
        app.logger.warning(f"estimate catalog seed skipped: {e}")

app.config["SESSION_PERMANENT"] = config.SESSION_PERMANENT
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)
app.config["UPLOAD_FOLDER"] = config.UPLOAD_FOLDER
app.config["SESSION_COOKIE_SECURE"] = True   # requires HTTPS in production
app.config["SESSION_COOKIE_HTTPONLY"] = True

limiter = Limiter(
    get_remote_address,
    app=app,
    storage_uri="memory://",
    default_limits=[],
)
# -------------------------------
# Flask-Mail and Login Configuration
# -------------------------------

# Mail settings are sourced from environment variables in ``config.py``.
# This avoids hard-coding credentials in the repository.
app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = "aliant.delgado07@gmail.com"
app.config["MAIL_PASSWORD"] = "lgco kmqe emqr qdrj"  # Use an app-specific password if using 2FA
app.config["MAIL_DEFAULT_SENDER"] = "aliant.delgado07@gmail.com"

mail = Mail(app)
serializer = URLSafeTimedSerializer(app.secret_key)


@app.errorhandler(429)
def ratelimit_handler(e):
    flash("Too many login attempts. Please wait and try again.", "danger")
    return redirect(url_for("login"))
# Buffer to temporarily store generated order summary PDFs
pdf_buffer: io.BytesIO | None = None
# -------------------------------
# Jinja Filters
# -------------------------------

@app.template_filter("datetimeformat")
def datetimeformat(value, fmt: str = "%Y-%m-%d %H:%M:%S"):
    """Format a timestamp for display in templates."""
    return datetime.fromtimestamp(value).strftime(fmt)
# -------------------------------
# Global Before-Request Handler (Session Timeout)
# -------------------------------

@app.before_request
def check_session_timeout():
    if request.endpoint in ('login', 'verify_code', 'verify_login', 'static'):
        return

    if "email" in session:
        # Backfill role for sessions created before role tracking was added
        if "role" not in session:
            user = get_user(session["email"])
            if user:
                session["role"] = user["role"]

        if session.get("email") == "zamoraplumbing01@gmail.com":
            return
        last_activity = session.get("last_activity", time.time())
        if time.time() - last_activity > 18000000:
            session.clear()
            flash("Session expired due to inactivity. Please log in again.", "warning")
            return redirect(url_for("login"))
        session["last_activity"] = time.time()

# -------------------------------
# Login Helper Functions
# -------------------------------

def is_logged_in():
    if "email" in session:
        # Always consider the Zamora Plumbing account as active.
        if session.get("email") == "zamoraplumbing01@gmail.com":
            return True
        last_activity = session.get("last_activity", time.time())
        if time.time() - last_activity > 18000000:
            session.pop("email", None)
            flash("Session expired due to inactivity. Please log in again.", "warning")
            return False
        session["last_activity"] = time.time()
        return True
    return False

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash("Please log in to access this page.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash("Please log in to access this page.", "warning")
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("Access denied.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated_function

SUPPLY_CODES = {
    "supply1": "BPS",
    "supply2": "S2",
    "supply3": "LPS",
    "supply4": "BOND",
}


def _ensure_supply_loaded(supply: str):
    loaders = {
        "supply1": du.load_default_file,
        "supply2": du.load_supply2_file,
        "supply3": du.load_supply3_file,
        "supply4": du.load_supply4_file,
    }
    loader = loaders.get(supply)
    if loader:
        loader()


def default_nav_links() -> list[dict[str, str]]:
    """Return the default navigation links for the React shell."""
    links = [
        {"label": "Home",          "href": url_for("index"),          "page": "home"},
        {"label": "View All",      "href": url_for("view_all"),        "page": "view_all"},
        {"label": "Search",        "href": url_for("search"),          "page": "search"},
        {"label": "Analyze",       "href": url_for("analyze"),         "page": "analyze"},
        {"label": "Material List", "href": url_for("material_list"),   "page": "material_list"},
        {"label": "Templates",     "href": url_for("templates_list"),  "page": "templates"},
        {"label": "Estimates",     "href": url_for("estimates_list"),  "page": "estimates"},
        {"label": "Upload PDF",    "href": url_for("upload_pdf"),      "page": "upload_pdf"},
    ]
    if session.get("role") == "admin":
        links.append({"label": "⚙ Admin", "href": url_for("admin_users"), "page": "admin"})
    return links


def render_app(
    page: str,
    initial_data: dict | None = None,
    nav_links: list[dict[str, str]] | None = None,
    status_code: int = 200,
):
    """Render the universal React shell template with serialized state."""

    context = {
        "page": page,
        "initial_data": initial_data or {},
        "nav_links": default_nav_links() if nav_links is None else nav_links,
        "flashes": get_flashed_messages(with_categories=True),
        "user_email": session.get("email"),
        "logout_url": url_for("logout") if session.get("email") else url_for("login"),
    }
    response = make_response(render_template("app.html", **context))
    response.status_code = status_code
    return response


def _search_supply_data(
    supply: str,
    query: str,
    page: int | None = None,
    per_page: int | None = None,
):
    """Search the in-memory catalog DataFrame and shape the result for JSON/React."""
    if not query:
        return {"rows": [], "columns": [], "next_page": None, "prev_page": None}

    df = get_catalog_df()
    if df is None or df.empty:
        return {"rows": [], "columns": [], "next_page": None, "prev_page": None}

    # Filter by supplier when not searching all
    code = SUPPLY_CODES.get(supply)
    if code:
        df = df[df["Supply"] == code]

    # Multi-keyword case-insensitive match on Description
    for kw in query.lower().split():
        df = df[df["Description"].str.lower().str.contains(kw, na=False, regex=False)]

    if df.empty:
        return {"rows": [], "columns": [], "next_page": None, "prev_page": None}

    df = df.sort_values(["Description", "Date"], ascending=[True, False])

    columns = ["Item Number", "Description", "Price per Unit", "Unit", "Invoice No.", "Date"]
    if supply == "all":
        columns.append("Supply")
    existing_cols = [c for c in columns if c in df.columns]

    rows = df[existing_cols].to_dict(orient="records")

    # Mark first 3 rows per description as recent; the rest are historical
    desc_counts: dict[str, int] = {}
    for row in rows:
        desc = row.get("Description", "")
        idx = desc_counts.get(desc, 0)
        row["is_recent"] = idx < 3
        desc_counts[desc] = idx + 1

    if supply != "all":
        for row in rows:
            desc = row.get("Description")
            if desc:
                row["graphUrl"] = url_for(
                    "product_detail", description=desc, supply=supply, ref="search", query=query
                )

    return {"rows": rows, "columns": existing_cols, "next_page": None, "prev_page": None}


def _analyze_price_changes(supply: str, start_date: str, end_date: str) -> dict:
    """Run the price change analysis and return JSON-serializable results."""
    _ensure_supply_loaded(supply)
    current_df = du.get_current_dataframe(supply)
    if current_df is None:
        return {"rows": [], "columns": []}

    try:
        start = pd.to_datetime(start_date)
        end = pd.to_datetime(end_date)
    except Exception:
        return {"rows": [], "columns": []}

    filtered = current_df[(current_df["Date"] >= start) & (current_df["Date"] <= end)]
    if filtered.empty:
        return {"rows": [], "columns": []}

    grouped = filtered.groupby(["Description", filtered["Date"].dt.to_period("M")])
    result_records: list[dict] = []
    for (desc, month), group in grouped:
        avg_price = group["Price per Unit"].mean()
        next_month = month + 1
        if (desc, next_month) not in grouped.groups:
            continue
        next_group = grouped.get_group((desc, next_month))
        next_avg_price = next_group["Price per Unit"].mean()
        if avg_price != next_avg_price:
            result_records.extend(group.to_dict("records"))
            result_records.extend(next_group.to_dict("records"))

    if not result_records:
        return {"rows": [], "columns": []}

    results_df = pd.DataFrame(result_records)
    if "Date" in results_df.columns:
        results_df["Date"] = pd.to_datetime(results_df["Date"]).dt.strftime("%Y-%m-%d")

    columns = [
        col
        for col in [
            "Description",
            "Item Number",
            "Price per Unit",
            "Unit",
            "Invoice No.",
            "Date",
        ]
        if col in results_df.columns
    ]
    shaped = results_df[columns] if columns else results_df
    return {"rows": shaped.to_dict(orient="records"), "columns": columns or list(results_df.columns)}


# ── Template helpers ──────────────────────────────────────────────────────────

def _split_template_path(full_name: str):
    """Return (folder, name) from 'folder/name' or 'name'."""
    if "/" in full_name:
        folder, name = full_name.rsplit("/", 1)
        return folder.strip(), name.strip()
    return "", full_name.strip()


def _template_entry(t: dict) -> dict:
    """Build the dict that the React app expects from a DB template row."""
    folder    = t.get("folder", "")
    name      = t["name"]
    full_name = f"{folder}/{name}" if folder else name
    try:
        mtime = datetime.strptime(t["updated_at"], "%Y-%m-%d %H:%M:%S").timestamp()
    except Exception:
        mtime = 0.0
    try:
        content   = json.loads(t["data"])
        products  = content.get("products", []) if isinstance(content, dict) else content
        item_count = len(products)
        subtotal  = 0.0
        for item in products:
            try:
                total = item.get("total")
                subtotal += float(total) if total is not None else (
                    float(item.get("last_price") or item.get("Last Price") or 0)
                    * float(item.get("quantity", 0))
                )
            except (TypeError, ValueError):
                pass
    except Exception:
        products, item_count, subtotal = [], 0, 0.0
    tax = subtotal * config.TAX_RATE
    return {
        "id":             t["id"],
        "name":           name,
        "full_name":      full_name,
        "group":          folder,
        "owner_email":    t.get("owner_email", ""),
        "mtime":          mtime,
        "item_count":     item_count,
        "total_with_tax": subtotal + tax,
        "edit_url":       url_for("edit_template",    name=full_name),
        "delete_url":     url_for("delete_template",  name=full_name),
        "rename_url":     url_for("rename_template",  name=full_name),
        "move_url":       url_for("move_template",    name=full_name),
        "versions_url":   url_for("api_template_versions", name=full_name),
    }


# (GitHub sync removed — templates are now stored in Turso)

# Load predefined material list DataFrames on startup
du.load_underground_list()
du.load_rough_list()
du.load_final_list()


# -------------------------------
# Routes for Main Functionality (Protected by login_required)
# -------------------------------

# Main Menu – note: the file upload functionality has been removed.
@app.route("/")
@login_required
def index():
    template_count = count_templates_db(session.get("email", ""), session.get("role", "user"))

    _df = get_catalog_df()
    def _unique_count(code):
        if _df is None or _df.empty:
            return 0
        return int(_df[_df["Supply"] == code]["Description"].nunique())

    stats = {
        "supply1Count": _unique_count("BPS"),
        "supply2Count": _unique_count("S2"),
        "supply3Count": _unique_count("LPS"),
        "supply4Count": _unique_count("BOND"),
        "templateCount": template_count,
    }

    initial = {
        "pageTitle": "Zamora Plumbing Corp Material Analyzer",
        "stats": stats,
        "actions": [
            {"label": "View All Content", "href": url_for("view_all"), "variant": "secondary"},
            {"label": "Search Description", "href": url_for("search"), "variant": "secondary"},
            {"label": "Analyze Price Changes", "href": url_for("analyze"), "variant": "secondary"},
            {"label": "Templates", "href": url_for("templates_list"), "variant": "secondary"},
            {"label": "Material List", "href": url_for("material_list"), "variant": "primary"},
            {"label": "Upload PDF", "href": url_for("upload_pdf"), "variant": "secondary"},
        ],
    }
    return render_app("home", initial)

@app.route("/view_all", methods=["GET"])
@login_required
def view_all():
    supply = request.args.get("supply", "supply1")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 200, type=int)

    code = SUPPLY_CODES.get(supply)
    df = get_catalog_df()

    if df is not None and not df.empty:
        if code:
            df = df[df["Supply"] == code]
        # Latest price per description: sort desc by Date, keep first occurrence
        df = (
            df.sort_values("Date", ascending=False)
            .drop_duplicates(subset=["Description"], keep="first")
            .sort_values("Description")
        )
    else:
        df = pd.DataFrame(columns=["Item Number", "Description", "Price per Unit", "Unit", "Date"])

    columns = ["Item Number", "Description", "Price per Unit", "Unit", "Date"]
    existing_cols = [c for c in columns if c in df.columns]
    total_rows = len(df)
    start = (page - 1) * per_page
    df_page = df.iloc[start : start + per_page]

    rows = []
    for record in df_page[existing_cols].to_dict(orient="records"):
        payload_row = {col: record.get(col, "") for col in existing_cols}
        desc = record.get("Description")
        if desc:
            payload_row["graphUrl"] = url_for(
                "product_detail", description=desc, supply=supply, ref="view_all"
            )
        rows.append(payload_row)

    supply_options = [
        {"value": "supply1", "label": "Supply 1"},
        {"value": "supply2", "label": "Supply 2"},
        {"value": "supply3", "label": "Lion Plumbing Supply"},
        {"value": "supply4", "label": "Bond Plumbing Supply"},
    ]

    payload = {
        "supply": supply,
        "columns": existing_cols,
        "rows": rows,
        "supplyOptions": supply_options,
        "productDetailBase": url_for("product_detail"),
        "viewAllUrl": url_for("view_all"),
        "page": page,
        "perPage": per_page,
        "totalRows": total_rows,
        "nextPage": page + 1 if (start + per_page) < total_rows else None,
        "prevPage": page - 1 if page > 1 else None,
    }

    if request.args.get("format") == "json":
        return jsonify(payload)

    return render_app("view_all", payload)

@app.route("/search", methods=["GET", "POST"])
@login_required
def search():
    """Render the React search page or redirect legacy POST submissions."""

    if request.method == "POST" and not request.is_json:
        supply_value = request.form.get("supply", "supply1")
        query_value = request.form.get("query", "").strip()
        if not query_value:
            flash("⚠ Please enter a search term.")
            return redirect(url_for("search", supply=supply_value))
        return redirect(url_for("search", supply=supply_value, query=query_value))

    supply = request.args.get("supply", "supply1")
    query = request.args.get("query", "")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", type=int)

    results_payload = _search_supply_data(supply, query, page, per_page)

    supply_options = [
        {"value": "supply1", "label": "Supply 1"},
        {"value": "supply2", "label": "Supply 2"},
        {"value": "supply3", "label": "Lion Plumbing Supply"},
        {"value": "supply4", "label": "Bond Plumbing Supply"},
        {"value": "all", "label": "All Supplies"},
    ]

    initial = {
        "supply": supply,
        "query": query,
        "columns": results_payload.get("columns", []),
        "rows": results_payload.get("rows", []),
        "nextPage": results_payload.get("next_page"),
        "prevPage": results_payload.get("prev_page"),
        "supplyOptions": supply_options,
        "searchApi": url_for("api_search"),
    }

    return render_app("search", initial)


@app.route("/api/search", methods=["GET"])
@login_required
def api_search():
    supply = request.args.get("supply", "supply1")
    query = request.args.get("query", "")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", type=int)
    payload = _search_supply_data(supply, query, page, per_page)
    return jsonify(
        {
            "columns": payload.get("columns", []),
            "rows": payload.get("rows", []),
            "next_page": payload.get("next_page"),
            "prev_page": payload.get("prev_page"),
        }
    )
    
@app.route("/api/sku_judge", methods=["POST"])
@login_required
def api_sku_judge():
    """
    JSON body: {"a": "<desc from supply A>", "b": "<desc from supply B>", "use_web": true}
    Returns JSON from judge_same_product.
    """
    try:
        data = request.get_json(force=True, silent=False) or {}
        a = (data.get("a") or "").strip()
        b = (data.get("b") or "").strip()
        use_web = bool(data.get("use_web", True))
        if not a or not b:
            return jsonify({"error": "Both 'a' and 'b' descriptions are required."}), 400

        app.logger.info(f"[sku_judge] Judging A='{a}' vs B='{b}', use_web={use_web}")
        result = judge_same_product(a, b, use_web=use_web, max_snippets=8)
        return jsonify(result), 200
    except Exception as e:
        app.logger.exception("SKU judge error")
        return jsonify({"error": str(e)}), 500

@app.route("/graph")
@login_required
def graph():
    """Generate a graph of Price per Unit over time for a given description from the selected supply."""
    supply = request.args.get("supply", "supply1")
    description = request.args.get("description")
    if not description:
        flash("⚠ Data or description not provided.")
        return redirect(url_for("index"))

    supplier_code = SUPPLY_CODES.get(supply)
    df = get_catalog_df()
    if df is None or df.empty:
        flash("⚠ No data available.")
        return redirect(url_for("index"))

    item_df = df[
        (df["Description"].str.lower() == description.lower()) &
        (df["Supply"] == supplier_code)
    ].copy()

    if item_df.empty:
        flash("⚠ No data available for the selected description.")
        return redirect(url_for("view_all", supply=supply))

    item_df = item_df.dropna(subset=["Date"]).sort_values(by="Date")
    fig, ax = plt.subplots(figsize=(8, 6))
    ax.plot(item_df["Date"], item_df["Price per Unit"], marker="o")
    ax.set_title(f"Prices Over Time for '{description}'")
    ax.set_xlabel("Date")
    ax.set_ylabel("Price per Unit")
    ax.grid(True)
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()

    output = io.BytesIO()
    FigureCanvas(fig).print_png(output)
    plt.close(fig)
    output.seek(0)
    return send_file(output, mimetype="image/png")

@app.route("/graph_data")
@login_required
def graph_data():
    supply = request.args.get("supply", "supply1")
    description = request.args.get("description")
    if not description:
        return jsonify({"dates": [], "prices": []})

    supplier_code = SUPPLY_CODES.get(supply)
    df = get_catalog_df()
    if df is None or df.empty:
        return jsonify({"dates": [], "prices": []})

    item_df = df[
        (df["Description"].str.lower() == description.lower()) &
        (df["Supply"] == supplier_code)
    ].copy()

    item_df = item_df.dropna(subset=["Date"]).sort_values(by="Date")
    dates = item_df["Date"].tolist()
    prices = item_df["Price per Unit"].tolist()
    return jsonify({"dates": dates, "prices": prices})

@app.route("/analyze", methods=["GET", "POST"])
@login_required
def analyze():
    """
    Analyze price changes for items across a custom date range from the selected supply.
    """
    supply = request.args.get("supply", "supply1")
    _ensure_supply_loaded(supply)
    current_df = du.get_current_dataframe(supply)
    if current_df is None:
        flash("⚠ Please ensure the Excel file for the selected supply is available.")
        return redirect(url_for("index"))

    if request.method == "POST" and request.is_json:
        data = request.get_json(force=True) or {}
        supply_value = data.get("supply", supply)
        start_date = data.get("start_date")
        end_date = data.get("end_date")
        result_payload = _analyze_price_changes(supply_value, start_date, end_date)
        return jsonify(result_payload)

    default_end = datetime.today().date()
    default_start = default_end - timedelta(days=30)

    if request.method == "POST":
        supply = request.form.get("supply", supply)
        start_date = request.form.get("start_date") or default_start.isoformat()
        end_date = request.form.get("end_date") or default_end.isoformat()
        result_payload = _analyze_price_changes(supply, start_date, end_date)
        if not result_payload.get("rows"):
            flash("⚠ No price changes found in the selected range.")
    else:
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        if start_date and end_date:
            result_payload = _analyze_price_changes(supply, start_date, end_date)
        else:
            start_date = default_start.isoformat()
            end_date = default_end.isoformat()
            result_payload = {"rows": [], "columns": []}

    supply_options = [
        {"value": "supply1", "label": "Supply 1"},
        {"value": "supply2", "label": "Supply 2"},
        {"value": "supply3", "label": "Lion Plumbing Supply"},
        {"value": "supply4", "label": "Bond Plumbing Supply"},
    ]

    initial = {
        "supply": supply,
        "startDate": start_date,
        "endDate": end_date,
        "columns": result_payload.get("columns", []),
        "rows": result_payload.get("rows", []),
        "supplyOptions": supply_options,
        "analyzeApi": url_for("analyze"),
    }

    return render_app("analyze", initial)

@app.route("/product_detail", methods=["GET"])
@login_required
def product_detail():
    """
    Display a page for a specific product (from the selected supply) with:
      - A graph (rendered by the /graph endpoint)
      - A table with Date and Price per Unit for that product (sorted by Date).
    """
    supply = request.args.get("supply", "supply1")
    description = request.args.get("description")
    if not description:
        flash("⚠ Please provide a product description.")
        return redirect(url_for("index"))

    supplier_code = SUPPLY_CODES.get(supply)
    df = get_catalog_df()
    if df is None or df.empty:
        flash("⚠ No catalog data available.")
        return redirect(url_for("index"))

    item_df = df[
        (df["Description"].str.lower() == description.lower()) &
        (df["Supply"] == supplier_code)
    ].copy()

    if item_df.empty:
        flash("⚠ No data available for the selected product.")
        return redirect(url_for("view_all", supply=supply))

    item_df = item_df.dropna(subset=["Date"]).sort_values(by="Date")
    dates = item_df["Date"].tolist()
    prices = item_df["Price per Unit"].tolist()

    ref = request.args.get("ref", "view_all")
    query = request.args.get("query", "")
    if ref == "search":
        back_url = url_for("search", supply=supply, query=query)
    else:
        back_url = url_for("view_all", supply=supply)

    initial = {
        "description": description,
        "supply": supply,
        "rows": [{"Date": d, "Price per Unit": p} for d, p in zip(dates, prices)],
        "chart": {"dates": dates, "prices": prices},
        "backUrl": back_url,
    }

    return render_app("product_detail", initial)

@app.route("/material_list", methods=["GET", "POST"])
@login_required
def material_list():
    if request.method == "POST":
        contractor = request.form.get("contractor")
        address = request.form.get("address")
        order_date = request.form.get("date")
        product_data_json = request.form.get("product_data")
        try:
            product_data = json.loads(product_data_json) if product_data_json else []
        except Exception as e:
            flash("Error processing product data.", "danger")
            return redirect(url_for("material_list"))
        
        # Retrieve include_price choice from the form:
        include_price = request.form.get("include_price", "yes")
        
        subtotal = sum(float(item.get("total", 0)) for item in product_data)
        tax = subtotal * 0.07
        total_cost = subtotal + tax

        # wkhtmltopdf/Qt requires XDG_RUNTIME_DIR to be owned by the current
        # user. /tmp is owned by root so we use a per-process temp dir instead.
        _xdg = tempfile.mkdtemp(prefix="wkhtml_")
        os.chmod(_xdg, stat.S_IRWXU)
        os.environ["XDG_RUNTIME_DIR"] = _xdg
        css_path = os.path.join(app.root_path, "static", "css", "order_summary.css")
        rendered = render_template(
            "order_summary.html",
            contractor=contractor,
            address=address,
            order_date=order_date,
            products=product_data,
            subtotal=subtotal,
            tax=tax,
            total_cost=total_cost,
            include_price=include_price,
            css_link=f"file://{css_path}",
        )
        import pdfkit
        try:
            options = {
                "enable-local-file-access": None,
                "margin-top":    "28",
                "margin-bottom": "22",
                "quiet":         "",
            }
            pdf = pdfkit.from_string(rendered, False, options=options, css=css_path)
            global pdf_buffer
            pdf_buffer = io.BytesIO(pdf)
            pdf_buffer.seek(0)
            # Persist PDF to a temporary file for reliability across workers
            old_path = session.pop("pdf_path", None)
            if old_path and os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except OSError:
                    pass
            address_filename = secure_filename(address) or "order_summary"
            pdf_filename = f"{address_filename}_{uuid.uuid4().hex}.pdf"
            pdf_path = os.path.join(app.config["UPLOAD_FOLDER"], pdf_filename)
            with open(pdf_path, "wb") as f:
                f.write(pdf)
            session["pdf_path"] = pdf_path
            session["last_address"] = address
            session["last_address_filename"] = address_filename
        except Exception as e:
            flash(f"PDF generation failed: {e}", "danger")
            return redirect(url_for("material_list"))

        recipient = session.get("email")
        subject_address = address or ""
        msg = Message(f"Material list- {subject_address}", recipients=[recipient])
        msg.body = "Please find attached your order summary PDF."
        msg.attach(f"{address_filename}.pdf", "application/pdf", pdf)
        try:
            mail.send(msg)
            flash("Order summary PDF sent to your email.", "success")
        except Exception as e:
            flash(f"Error sending email: {e}", "danger")

        pdf_buffer.seek(0)
        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{address_filename}.pdf",
        )
    
    # For GET: load predetermined or saved templates
    list_option = request.args.get("list", "underground")
    viewer_email = session.get("email", "")
    viewer_role  = session.get("role", "user")
    db_templates = list_templates_db(viewer_email, viewer_role)
    custom_templates = {}
    template_folders = set()
    for t in db_templates:
        folder    = t.get("folder", "")
        tname     = t["name"]
        full_name = f"{folder}/{tname}" if folder else tname
        if folder:
            template_folders.add(folder)
        try:
            custom_templates[full_name] = json.loads(t["data"])
        except Exception:
            pass
    list_option_lower = list_option.lower()
    project_info = {"contractor": "", "address": "", "date": ""}
    if list_option in custom_templates:
        raw_template = custom_templates[list_option]
        if isinstance(raw_template, dict):
            project_info = raw_template.get("project_info", project_info)
            raw_list = raw_template.get("products", [])
        else:
            raw_list = raw_template
        product_list = []
        for item in raw_list:
            if item.get("type") == "divider":
                product_list.append({"type": "divider", "label": item.get("label", "")})
                continue
            desc = (
                item.get("Product Description")
                or item.get("description")
                or item.get("Description")
                or ""
            )
            price = (
                item.get("Last Price")
                or item.get("last_price")
                or item.get("Price per Unit")
                or 0
            )
            qty = item.get("quantity", 0)
            unit = item.get("Unit") or item.get("unit", "")
            total = item.get("total") or item.get("Total") or 0
            supply = item.get("Supply") or item.get("supply", "BPS")
            product_list.append(
                {
                    "Product Description": desc,
                    "Last Price": price,
                    "quantity": qty,
                    "Unit": unit,
                    "total": total,
                    "Supply": supply,
                }
            )
    elif list_option_lower == "underground":
        du.update_underground_prices()
        product_list = du.df_underground.to_dict("records") if du.df_underground is not None else []
    elif list_option_lower == "rough":
        du.update_rough_prices()
        product_list = du.df_rough.to_dict("records") if du.df_rough is not None else []
    elif list_option_lower == "final":
        du.update_final_prices()
        product_list = du.df_final.to_dict("records") if du.df_final is not None else []
    elif list_option_lower == "new":
        product_list = []
    else:
        product_list = []

    _cat = get_catalog_df()
    def _catalog_for(code):
        if _cat is None or _cat.empty:
            return []
        sub = _cat[_cat["Supply"] == code]
        return (
            sub.sort_values("Date", ascending=False)
            .drop_duplicates(subset=["Description"], keep="first")
            .sort_values("Description")
            .to_dict(orient="records")
        )

    supply1_products = _catalog_for("BPS")
    supply2_products = _catalog_for("S2")
    supply3_products = _catalog_for("LPS")
    supply4_products = _catalog_for("BOND")
    template_name_arg = request.args.get("template_name", "")
    if not template_name_arg and list_option_lower not in ["underground", "rough", "final", "new"]:
        template_name_arg = list_option
    full_template_name = template_name_arg
    template_folder = ""
    template_name = template_name_arg
    if template_name_arg:
        template_folder, template_name = os.path.split(template_name_arg)
        if template_folder == ".":
            template_folder = ""
    initial = {
        "listOption": list_option,
        "products": product_list,
        "projectInfo": project_info,
        "customTemplates": sorted(custom_templates.keys()),
        "templateFolders": sorted(template_folders),
        "templateName": template_name,
        "templateFolder": template_folder,
        "fullTemplateName": full_template_name,
        "catalog": {
            "supply1": supply1_products,
            "supply2": supply2_products,
            "supply3": supply3_products,
            "supply4": supply4_products,
        },
        "listUrl": url_for("material_list"),
        "downloadUrl": url_for("download_summary"),
        "saveTemplateUrl": url_for("save_template"),
    }

    return render_app("material_list", initial)

@app.route("/save_template", methods=["POST"])
@login_required
def save_template():
    template_name_full = request.form.get("template_name", "").strip()
    product_data       = request.form.get("product_data")
    project_info_str   = request.form.get("project_info")

    if not template_name_full or not product_data:
        flash("Template name and data are required.", "danger")
        return redirect(url_for("material_list"))

    folder, tname = _split_template_path(template_name_full)

    try:
        products = json.loads(product_data)
    except Exception:
        flash("Invalid product data.", "danger")
        return redirect(url_for("material_list"))
    try:
        project_info = json.loads(project_info_str) if project_info_str else {}
    except Exception:
        project_info = {}

    data_json = json.dumps({"project_info": project_info, "products": products})
    actor_email = session["email"]
    actor_role  = session.get("role", "user")

    tid = save_template_db(tname, folder, actor_email, actor_role, data_json)
    if tid is None:
        flash("Access denied: a template with that name belongs to another user.", "danger")
    else:
        flash("Template saved.", "success")
    return redirect(url_for("material_list"))

# -------------------------------
# Template Management Routes
# -------------------------------

@app.route("/templates")
@login_required
def templates_list():
    sort_key     = request.args.get("sort", "name")
    group_by     = request.args.get("group", "folder")
    viewer_email = session.get("email", "")
    viewer_role  = session.get("role", "user")

    db_rows = list_templates_db(viewer_email, viewer_role)
    entries = [_template_entry(t) for t in db_rows]
    folders = sorted({e["group"] for e in entries if e["group"]})

    if sort_key == "date":
        entries.sort(key=lambda x: x["mtime"], reverse=True)
    else:
        entries.sort(key=lambda x: x["name"].lower())

    grouped = {}
    if group_by == "folder":
        for e in entries:
            grouped.setdefault(e["group"], []).append(e)

    initial = {
        "entries":          entries,
        "grouped":          grouped if group_by == "folder" else None,
        "sortKey":          sort_key,
        "groupBy":          group_by,
        "folders":          folders,
        "createFolderUrl":  url_for("create_template_folder"),
        "isAdmin":          viewer_role == "admin",
    }
    return render_app("templates", initial)


@app.route("/edit_template/<path:name>")
@login_required
def edit_template(name):
    return redirect(url_for("material_list", list=name, template_name=name))


@app.route("/api/template_preview")
@login_required
def api_template_preview():
    full_name = request.args.get("name", "").strip()
    if not full_name:
        return jsonify({"products": []}), 400
    folder, tname = _split_template_path(full_name)
    t = get_template_db(tname, folder, session.get("email", ""), session.get("role", "user"))
    if not t:
        return jsonify({"products": []}), 404
    try:
        content  = json.loads(t["data"])
        products = content.get("products", []) if isinstance(content, dict) else content
        return jsonify({"products": products})
    except Exception as e:
        app.logger.error(f"Template preview error: {e}")
        return jsonify({"products": [], "error": str(e)}), 500


@app.route("/api/duplicate_template", methods=["POST"])
@login_required
def api_duplicate_template():
    source_name = request.form.get("source_name", "").strip()
    new_name    = request.form.get("new_name", "").strip()
    if not source_name or not new_name:
        return jsonify({"error": "source_name and new_name are required"}), 400
    src_folder, src_tname = _split_template_path(source_name)
    dst_folder, dst_tname = _split_template_path(new_name)
    actor_email = session["email"]
    actor_role  = session.get("role", "user")
    ok = duplicate_template_db(src_tname, src_folder, dst_tname, dst_folder, actor_email, actor_role)
    if not ok:
        return jsonify({"error": "Source not found, destination already exists, or access denied."}), 409
    return jsonify({"success": True, "new_name": new_name})


@app.route("/delete_template/<path:name>", methods=["POST"])
@login_required
def delete_template(name):
    folder, tname = _split_template_path(name)
    ok = delete_template_db(tname, folder, session["email"], session.get("role", "user"))
    if ok:
        flash("Template deleted.", "info")
    else:
        flash("Template not found or access denied.", "danger")
    return redirect(request.referrer or url_for("templates_list"))


@app.route("/create_template_folder", methods=["POST"])
@login_required
def create_template_folder():
    # Folders are virtual in the DB (just the folder column).
    # They appear automatically when a template is saved into them.
    folder_name = request.form.get("folder_name", "").strip()
    if folder_name:
        flash(f'Folder "{folder_name}" ready — save a template into it to make it appear.', "info")
    return redirect(request.referrer or url_for("templates_list"))


@app.route("/rename_template/<path:name>", methods=["POST"])
@login_required
def rename_template(name):
    new_name = request.form.get("new_name", "").strip()
    if not new_name:
        flash("New name required.", "danger")
        return redirect(request.referrer or url_for("templates_list"))
    old_folder, old_tname = _split_template_path(name)
    new_folder, new_tname = _split_template_path(new_name)
    ok = rename_template_db(old_tname, old_folder, new_tname, new_folder,
                            session["email"], session.get("role", "user"))
    if ok:
        flash("Template renamed.", "success")
    else:
        flash("Rename failed: template not found, name already taken, or access denied.", "danger")
    return redirect(url_for("templates_list"))


@app.route("/move_template/<path:name>", methods=["POST"])
@login_required
def move_template(name):
    target_folder = request.form.get("target_folder", "").strip()
    new_folder    = request.form.get("new_folder", "").strip()
    destination   = new_folder or target_folder
    old_folder, tname = _split_template_path(name)
    ok = rename_template_db(tname, old_folder, tname, destination,
                            session["email"], session.get("role", "user"))
    if ok:
        flash("Template moved.", "success")
    else:
        flash("Move failed: template not found, destination exists, or access denied.", "danger")
    return redirect(url_for("templates_list"))


@app.route("/api/template_versions/<path:name>")
@login_required
def api_template_versions(name):
    folder, tname = _split_template_path(name)
    versions = get_template_versions_db(tname, folder, session.get("email", ""), session.get("role", "user"))
    return jsonify({"versions": versions})


@app.route("/restore_template_version/<int:version_id>", methods=["POST"])
@login_required
def restore_template_version(version_id):
    ok = restore_template_version_db(version_id, session["email"], session.get("role", "user"))
    if ok:
        flash("Template restored to selected version.", "success")
    else:
        flash("Restore failed: version not found or access denied.", "danger")
    return redirect(request.referrer or url_for("templates_list"))


# ── Estimate helpers & routes ─────────────────────────────────────────────────

def _row(qty, desc, unit_cost, comments="", add_comments=""):
    total = round((qty or 0) * (unit_cost or 0), 2)
    return {
        "id": str(uuid.uuid4()), "type": "manual",
        "qty": qty if qty is not None else "",
        "description": desc, "unit_cost": unit_cost or 0,
        "total": total, "comments": comments, "add_comments": add_comments,
    }


_TEMPLATE_SECTIONS = [
    ("Material Take Out", False, [
        _row(112, "Residential", 1000, "Waste and water. Drain includes the underground, rough in and final portion of the fixtures. These include material, labor and installation."),
        _row(None, "Floor drains for the ADA Bathroom (Not in plans)", 900, "Piping, furnish and installation of the FD + Trap primer"),
        _row(None, "Outdoor shower", 1000, "Water and installation"),
        _row(6, "Commercial Outlet", 1000, "Waste and water. Drain includes the underground, rough in and final portion of the fixtures. These include material, labor and installation."),
    ]),
    ("Water Distribution System", False, [
        _row(15, "Electrical tank water heater 30 gal", 1000, "Installation, pan drain and relief lines connection"),
        _row(None, "Electrical tankless water heater", 800, "Installation, relief line"),
        _row(None, "Hose Bibbs on the roof (standard) with SOV", 800, "Water piping + SOV+HB installation and labor"),
        _row(1, "Hose bibbs on dumpster room (standard) with SOV", 800, "Water piping + SOV+HB installation and labor"),
        _row(None, "Hydrant box BBQ area with SOV", 800, "Furnish, install and labor (model # 5509QTSAP SMITH WALL HYDRANT W/SS BOX)(400.00 Each)"),
        _row(None, "Hydrant box pool area (Below outdoor shower) with SOV", 800, "Furnish, install and labor (model # 5509QTSAP SMITH WALL HYDRANT W/SS BOX)"),
        _row(None, "Hydrant box for bathrooms", 800, "Furnish, install and labor (model # 5509QTSAP SMITH WALL HYDRANT W/SS BOX)"),
        _row(14, "Ice makers", 130, "Furnish and install ice maker box", "Fire rated?"),
        _row(None, "Hammer arrestors", 4000, "Furnish and install"),
        _row(None, "Mixing valve for lavatories", 0, "Furnish and install"),
        _row(1, "Pan drain Riser and horizontal Size CPVC", 4000, "Above and below grade to discharge above ground. Material and labor"),
        _row(1, "Pan Drain Connection", 10, "Connection to the pan drain"),
        _row(None, "Water distribution inside the unit", 0, "Material and labor"),
        _row(None, '1-1/4" CPVC Above Ceiling', 30, "Material and labor for the installation of the piping"),
        _row(200, '1-1/2" CPVC Above Ceiling', 40, "Material and labor for the installation of the piping"),
        _row(60, '2" CPVC above Ceiling', 50, "Material and labor for the installation of the piping"),
        _row(None, '2 1/2" CPVC above Ceiling', 51, "Material and labor for the installation of the piping"),
        _row(None, '3" CPVC Schedule 80 Above Ceiling', 70, "Material and labor for the installation of the piping"),
        _row(7, 'Water Riser Type 1"', 150, "Material and labor for the installation of the piping"),
        _row(None, 'Water Riser Type 1-1/4"', 10, "Material and labor for the installation of the piping"),
        _row(None, 'Water Riser Type 1-1/2"', 0, ""),
        _row(None, "Pool equipment 1-1/4 water and furnish and install 1-1/4 backflow preventor. With SOV", 2000, ""),
        _row(None, '3/4 S.O.V for 1 bath units (two per units)', 30, "Bronze or CPVC — decide"),
        _row(14, '1" S.O.V for 2 bath units (two per units)', 40, "Bronze or CPVC — decide"),
        _row(None, "SOV for the riser 1-1/2 CPVC SOV", 50, ""),
        _row(None, "Water submeter", 100, ""),
        _row(1, "Water service connection", 2000, ""),
        _row(None, '3" backflow preventor', 6000, ""),
        _row(None, "Water for the trash shutter", 1500, ""),
        _row(None, "Booster pump furnish", 35000, ""),
        _row(None, "Booster pump installation", 20000, ""),
    ]),
    ("Drainage System", False, [
        _row(1, "Sewer and connection", 2000, "Piping / connection / clean outs"),
        _row(400, '4" PVC Underground', 20, "Material and installation"),
        _row(None, '6" PVC Underground', 38, "Material and installation"),
        _row(None, '8" PVC Underground', 48, "Material and installation"),
        _row(None, 'PVC Stack up 3"', 5, "Riser, hanger, labor etc"),
        _row(None, 'Relief vent 3"', 5, "Riser, hanger, labor etc"),
        _row(None, 'Stack up from garage 3"', 5, "Riser, hanger, labor etc"),
        _row(None, 'Insulation for the san riser 3"', 5, "Material and labor"),
        _row(None, '4" PVC Underground Hanger', 15, "Teardrop hanger and asphalts"),
        _row(None, '6" PVC Underground Hanger', 15, "Teardrop hanger and asphalts"),
        _row(None, '8" PVC Underground Hanger', 15, "Teardrop hanger and asphalts"),
        _row(1, "Sump pump Installation", 2500, ""),
        _row(1, "Furnish sump pump", 4000, ""),
        _row(None, "Hub drain for the sump pump", 1500, ""),
        _row(None, "FD to the trash room", 1500, ""),
        _row(None, "Hub drain for the fire pump", 1500, ""),
        _row(None, "Lint intercept furnish", 2000, ""),
        _row(None, "Installation of lint intercept", 9000, ""),
    ]),
    ("A/C System", False, [
        _row(14, "Inside the units CPVC", 600, "UNDERGROUND A/C ONLY"),
        _row(None, '3" Underground line 200x2 missing', 20, "UNDERGROUND A/C ONLY"),
        _row(None, '2" Riser PVC', 10, "AIR HANDLER CLOSET Branch"),
        _row(None, '2" Above ground horizontal', 10, "AIR HANDLER CLOSET Branch"),
        _row(None, "Connection to french drain or Civil", 1000, ""),
        _row(None, '2" insulation 1/2 thickness', 5, ""),
        _row(None, '3" PVC Underground Hanger', 15, "Teardrop hanger and asphalts"),
    ]),
    ("Rain Water System", False, [
        _row(None, 'ROOF AREA DRAINS (4")', 700, "Furnish and install (250 each)", "Model: 21504-22-Y Cast iron JOSAM"),
        _row(None, 'AD (Deck Drain 4")', 800, "Furnish and install", "Per BH princeton"),
        _row(None, 'Planter Drain 3"', 1500, "Furnish and install"),
        _row(None, '3" PVC Underground', 0, "Material and installation"),
        _row(None, '4" PVC Underground', 20, "Material and installation"),
        _row(None, '6" PVC Underground', 38, "Material and installation"),
        _row(None, '8" PVC Underground', 48, "Material and installation"),
        _row(None, "SD Riser roof 45' 27 Riserx 45' (4\")", 20, "Riser, hanger, labor etc"),
        _row(None, "AD deck drain riser 11 riser x 16' (4\")", 20, "Riser, hanger, labor etc"),
        _row(None, "Planter drain riser pool deck 4x16' (3\")", 10, "Riser, hanger, labor etc"),
        _row(None, "Planter drain to civil", 3000, "Riser, hanger, labor etc"),
        _row(None, "Connection to french drain or Civil", 500, ""),
        _row(None, '4" Riser insulation (stacks only)', 10, "", "Verify material: fiberglass or armaflex"),
        _row(None, '4" horizontal ceiling space insulation', 10, "", "Verify material: fiberglass or armaflex"),
        _row(None, '4" PVC Underground Hanger', 15, "Teardrop hanger and asphalts"),
        _row(None, '6" PVC Underground Hanger', 15, "Teardrop hanger and asphalts"),
        _row(None, '8" PVC Underground Hanger', 15, "Teardrop hanger and asphalts"),
    ]),
    ("Extras", False, [
        _row(None, "Backfilling", 4000, ""),
        _row(1, "Sleeves", 2500, ""),
        _row(None, "Fixtures handling", 4000, ""),
        _row(None, "Excavation", 4000, ""),
        _row(None, "Sand", 3000, ""),
        _row(None, "Temporary works", 1500, ""),
    ]),
    ("Gas System", True, [
        _row(0, "Gas Outlets", 400, "Propane / Metalic pipe A / Flex B", "Gas not included in plumbing price"),
        _row(0, "Piping length", 20, "Furnish / installation / meter connection / MP"),
        _row(0, "Permit", 500, "License"),
        _row(0, "Extra", 1500, "Extra"),
        _row(0, "Tank 1000 gal.", 12000, "Furnish and install / concrete slab / no tank fill"),
    ]),
]


def _estimate_entry(e: dict) -> dict:
    folder    = e.get("folder", "")
    name      = e["name"]
    full_name = f"{folder}/{name}" if folder else name
    try:
        mtime = datetime.strptime(e["updated_at"], "%Y-%m-%d %H:%M:%S").timestamp()
    except Exception:
        mtime = 0.0
    try:
        content  = json.loads(e["data"])
        sections = content.get("sections", [])
        row_count = sum(len(s.get("rows", [])) for s in sections)
        plumbing_total = content.get("plumbing_total", 0.0)
        gas_total      = content.get("gas_total", 0.0)
        grand_total    = content.get("grand_total", plumbing_total)
    except Exception:
        row_count, plumbing_total, gas_total, grand_total = 0, 0.0, 0.0, 0.0
    return {
        "id":             e["id"],
        "name":           name,
        "full_name":      full_name,
        "group":          folder,
        "owner_email":    e.get("owner_email", ""),
        "mtime":          mtime,
        "row_count":      row_count,
        "plumbing_total": plumbing_total,
        "gas_total":      gas_total,
        "grand_total":    grand_total,
    }


def _build_empty_estimate():
    """Pre-populated template rows, all quantities cleared."""
    sections = [
        {
            "id": str(uuid.uuid4()), "name": name, "is_gas": is_gas,
            "rows": [{**r, "qty": "", "total": 0} for r in rows],
        }
        for name, is_gas, rows in _TEMPLATE_SECTIONS
    ]
    return {
        "project_info":   {"name": "", "address": "", "contractor": "", "date": ""},
        "sections":       sections,
        "plumbing_total": 0.0,
        "gas_total":      0.0,
        "grand_total":    0.0,
    }


def _build_blank_estimate():
    """Completely blank estimate — sections exist but no rows."""
    sections = [
        {"id": str(uuid.uuid4()), "name": name, "is_gas": is_gas, "rows": []}
        for name, is_gas, _ in _TEMPLATE_SECTIONS
    ]
    return {
        "project_info":   {"name": "", "address": "", "contractor": "", "date": ""},
        "sections":       sections,
        "plumbing_total": 0.0,
        "gas_total":      0.0,
        "grand_total":    0.0,
    }


# Seed catalog with template rows now that _TEMPLATE_SECTIONS is defined
_seed_estimate_catalog()


@app.route("/estimates")
@login_required
def estimates_list():
    raw = list_estimates_db(session.get("email", ""), session.get("role", "user"))
    entries = [_estimate_entry(e) for e in raw]
    return render_app("estimates", {
        "estimates":    entries,
        "newUrl":       url_for("estimate_builder"),
        "blankUrl":     url_for("estimate_builder", blank=1),
        "deleteUrl":    "/delete_estimate/",
        "duplicateUrl": url_for("api_duplicate_estimate"),
    })


@app.route("/estimate")
@login_required
def estimate_builder():
    name  = request.args.get("name", "").strip()
    blank = request.args.get("blank", "")
    if name:
        folder, ename = _split_template_path(name)
        e = get_estimate_db(ename, folder, session.get("email", ""), session.get("role", "user"))
        if e:
            content = json.loads(e["data"])
        else:
            flash("Estimate not found.", "warning")
            content = _build_empty_estimate()
    elif blank:
        content = _build_blank_estimate()
    else:
        content = _build_empty_estimate()

    # Attach available material list names for the link-ML modal
    ml_raw = list_templates_db(session.get("email", ""), session.get("role", "user"))
    ml_names = [
        (f"{t['folder']}/{t['name']}" if t.get("folder") else t["name"])
        for t in ml_raw
    ]

    # Refresh ML totals for any linked rows so edits to a ML propagate automatically
    for section in content.get("sections", []):
        for row in section.get("rows", []):
            if row.get("type") == "material_list" and row.get("material_list_name"):
                ml_name = row["material_list_name"]
                ml_folder, ml_tname = _split_template_path(ml_name)
                current_total = get_material_list_total(
                    ml_tname, ml_folder,
                    session.get("email", ""), session.get("role", "user")
                )
                if current_total is not None:
                    row["unit_cost"] = current_total
                    row["total"]     = current_total

    return render_app("estimate_builder", {
        "estimateName": name,
        "content":      content,
        "saveUrl":      url_for("save_estimate"),
        "exportPdfUrl": url_for("export_estimate_pdf"),
        "exportXlsUrl": url_for("export_estimate_excel"),
        "catalogUrl":   url_for("api_estimate_catalog"),
        "mlTotalUrl":   url_for("api_material_list_total"),
        "mlNames":      ml_names,
        "mlListUrl":    url_for("material_list"),
    })


@app.route("/save_estimate", methods=["POST"])
@login_required
def save_estimate():
    estimate_name = request.form.get("estimate_name", "").strip()
    data_json     = request.form.get("estimate_data", "")
    if not estimate_name or not data_json:
        return jsonify({"ok": False, "error": "Missing name or data"}), 400

    folder, ename = _split_template_path(estimate_name)

    # Keep totals up to date
    try:
        content = json.loads(data_json)
        plumbing_total = sum(
            float(r.get("total") or 0)
            for s in content.get("sections", []) if not s.get("is_gas")
            for r in s.get("rows", [])
        )
        gas_total = sum(
            float(r.get("total") or 0)
            for s in content.get("sections", []) if s.get("is_gas")
            for r in s.get("rows", [])
        )
        content["plumbing_total"] = round(plumbing_total, 2)
        content["gas_total"]      = round(gas_total, 2)
        content["grand_total"]    = round(plumbing_total + gas_total, 2)
        data_json = json.dumps(content)
    except Exception:
        pass

    # Auto-save catalog entries for manual rows
    try:
        display_name = f"{folder}/{ename}" if folder else ename
        for section in content.get("sections", []):
            for row in section.get("rows", []):
                if row.get("type") == "manual" and row.get("description", "").strip():
                    upsert_estimate_catalog(
                        description   = row["description"].strip(),
                        unit_cost     = float(row.get("unit_cost") or 0),
                        comments      = row.get("comments", ""),
                        add_comments  = row.get("add_comments", ""),
                        category      = section.get("name", ""),
                        estimate_name = display_name,
                    )
    except Exception:
        pass

    tid = save_estimate_db(ename, folder, session["email"], session.get("role", "user"), data_json)
    if tid is None:
        return jsonify({"ok": False, "error": "Access denied"}), 403
    return jsonify({"ok": True, "id": tid})


@app.route("/delete_estimate/<path:name>", methods=["POST"])
@login_required
def delete_estimate(name):
    folder, ename = _split_template_path(name)
    ok = delete_estimate_db(ename, folder, session["email"], session.get("role", "user"))
    if ok:
        flash("Estimate deleted.", "success")
    else:
        flash("Delete failed.", "danger")
    return redirect(url_for("estimates_list"))


@app.route("/api/duplicate_estimate", methods=["POST"])
@login_required
def api_duplicate_estimate():
    src  = request.form.get("src_name", "").strip()
    dst  = request.form.get("dst_name", "").strip()
    if not src or not dst:
        flash("Source and destination names required.", "danger")
        return redirect(url_for("estimates_list"))
    sf, sn = _split_template_path(src)
    df, dn = _split_template_path(dst)
    ok = duplicate_estimate_db(sn, sf, dn, df, session["email"], session.get("role", "user"))
    if ok:
        flash(f"Estimate duplicated as '{dst}'.", "success")
    else:
        flash("Duplicate failed: destination already exists or access denied.", "danger")
    return redirect(url_for("estimates_list"))


@app.route("/api/estimate_catalog")
@login_required
def api_estimate_catalog():
    q = request.args.get("q", "").strip()
    results = search_estimate_catalog(q, limit=20) if q else []
    return jsonify(results)


@app.route("/api/material_list_total")
@login_required
def api_material_list_total():
    name = request.args.get("name", "").strip()
    if not name:
        return jsonify({"ok": False}), 400
    folder, tname = _split_template_path(name)
    total = get_material_list_total(tname, folder, session.get("email", ""), session.get("role", "user"))
    if total is None:
        return jsonify({"ok": False, "error": "Not found"}), 404
    return jsonify({"ok": True, "total": total})


@app.route("/estimate/export_pdf", methods=["POST"])
@login_required
def export_estimate_pdf():
    import pdfkit
    data_json = request.form.get("estimate_data", "")
    try:
        content = json.loads(data_json)
    except Exception:
        flash("Invalid estimate data.", "danger")
        return redirect(url_for("estimates_list"))

    all_sections   = content.get("sections", [])
    plumb_sections = [s for s in all_sections if not s.get("is_gas")]
    gas_sections   = [s for s in all_sections if s.get("is_gas")]
    css_path = os.path.join(app.root_path, "static", "css", "order_summary.css")
    rendered = render_template(
        "estimate_summary.html",
        project_info    = content.get("project_info", {}),
        plumb_sections  = plumb_sections,
        gas_sections    = gas_sections,
        plumbing_total  = content.get("plumbing_total", 0.0),
        gas_total       = content.get("gas_total", 0.0),
        grand_total     = content.get("grand_total", 0.0),
        bids            = content.get("bids", []),
        css_link        = f"file://{css_path}",
    )

    try:
        pdf_bytes = pdfkit.from_string(rendered, False,
                                       options={"enable-local-file-access": None},
                                       css=css_path)
    except Exception as e:
        app.logger.error(f"pdfkit error: {e}")
        flash("PDF generation failed. Is wkhtmltopdf installed?", "danger")
        return redirect(url_for("estimates_list"))

    proj_name = content.get("project_info", {}).get("name", "estimate")
    safe_name = "".join(c for c in proj_name if c.isalnum() or c in " _-").strip() or "estimate"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"{safe_name}.pdf",
    )


@app.route("/estimate/export_excel", methods=["POST"])
@login_required
def export_estimate_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data_json = request.form.get("estimate_data", "")
    try:
        content = json.loads(data_json)
    except Exception:
        flash("Invalid estimate data.", "danger")
        return redirect(url_for("estimates_list"))

    pi = content.get("project_info", {})
    sections = content.get("sections", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate"

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 35

    header_fill   = PatternFill("solid", fgColor="1E3A5F")
    section_fill  = PatternFill("solid", fgColor="2B6CB0")
    subtotal_fill = PatternFill("solid", fgColor="EBF4FF")
    alt_fill      = PatternFill("solid", fgColor="F7FAFC")
    bold_white    = Font(bold=True, color="FFFFFF")
    bold_dark     = Font(bold=True, color="1E3A5F")
    thin = Side(style="thin", color="CBD5E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Project info header
    ws.merge_cells("A1:F1")
    ws["A1"] = "Zamora Plumbing Corp — Estimate"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("Project", pi.get("name", "")),
        ("Address", pi.get("address", "")),
        ("Contractor", pi.get("contractor", "")),
        ("Date", pi.get("date", "")),
    ]
    r = 2
    for label, val in info_rows:
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, val)
        r += 1

    r += 1  # blank row

    # Column headers
    headers = ["QTY", "SYSTEM / DESCRIPTION", "UNIT OUTLET COST", "TOTAL COST", "COMMENTS", "ADDITIONAL COMMENTS"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(r, col, h)
        c.font = bold_white
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    r += 1

    gas_fill = PatternFill("solid", fgColor="1A4731")

    def _write_section(ws, section, r, section_fill, alt_fill, border, bold_white, bold_dark, subtotal_fill):
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(r, 1, section.get("name", "").upper())
        c.font = bold_white
        c.fill = section_fill
        c.alignment = Alignment(horizontal="left")
        r += 1
        subtotal = 0.0
        for i, row in enumerate(section.get("rows", [])):
            fill = alt_fill if i % 2 == 0 else PatternFill()
            desc = row.get("description", "")
            if row.get("type") == "material_list":
                desc = f"[ML] {desc}"
            cells_data = [
                row.get("qty") if row.get("qty") != "" else "",
                desc,
                row.get("unit_cost") or 0,
                row.get("total") or 0,
                row.get("comments", ""),
                row.get("add_comments", ""),
            ]
            for col, val in enumerate(cells_data, 1):
                c = ws.cell(r, col, val)
                c.fill = fill
                c.border = border
                if col in (3, 4) and isinstance(val, (int, float)):
                    c.number_format = '"$"#,##0.00'
                if col == 2:
                    c.alignment = Alignment(wrap_text=True)
            subtotal += float(row.get("total") or 0)
            r += 1
        ws.merge_cells(f"A{r}:C{r}")
        ws.cell(r, 1, "Subtotal").font = bold_dark
        ws.cell(r, 1).fill = subtotal_fill
        ws.cell(r, 1).alignment = Alignment(horizontal="right")
        c = ws.cell(r, 4, subtotal)
        c.font = bold_dark
        c.fill = subtotal_fill
        c.number_format = '"$"#,##0.00'
        r += 1
        return r

    plumb_sections = [s for s in sections if not s.get("is_gas")]
    gas_sections   = [s for s in sections if s.get("is_gas")]

    for section in plumb_sections:
        r = _write_section(ws, section, r, section_fill, alt_fill, border, bold_white, bold_dark, subtotal_fill)

    # Plumbing Project Total
    r += 1
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(r, 1, "PLUMBING PROJECT TOTAL").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(r, 1).fill = header_fill
    ws.cell(r, 1).alignment = Alignment(horizontal="right")
    c = ws.cell(r, 4, content.get("plumbing_total", 0.0))
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = header_fill
    c.number_format = '"$"#,##0.00'
    r += 2

    # Gas System sections
    for section in gas_sections:
        r = _write_section(ws, section, r, gas_fill, alt_fill, border, bold_white, bold_dark, subtotal_fill)

    # Gas Total
    r += 1
    ws.merge_cells(f"A{r}:C{r}")
    ws.cell(r, 1, "GAS SYSTEM TOTAL").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(r, 1).fill = gas_fill
    ws.cell(r, 1).alignment = Alignment(horizontal="right")
    c = ws.cell(r, 4, content.get("gas_total", 0.0))
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = gas_fill
    c.number_format = '"$"#,##0.00'

    # Bids section
    bids = content.get("bids", [])
    r += 2
    bids_header_fill = PatternFill("solid", fgColor="374151")
    bids_col_fill    = PatternFill("solid", fgColor="4B5563")
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1, "BIDS")
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = bids_header_fill
    c.alignment = Alignment(horizontal="center")
    r += 1
    for col, label in [(1, "BID #"), (2, "AMOUNT ($)"), (3, "COMMENTS")]:
        c = ws.cell(r, col, label)
        c.font = bold_white
        c.fill = bids_col_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
    ws.merge_cells(f"C{r}:F{r}")
    r += 1
    for i, bid in enumerate(bids):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        c = ws.cell(r, 1, bid.get("bid_num", ""))
        c.fill = fill; c.border = border
        c = ws.cell(r, 2, float(bid.get("amount") or 0))
        c.fill = fill; c.border = border; c.number_format = '"$"#,##0.00'
        ws.merge_cells(f"C{r}:F{r}")
        c = ws.cell(r, 3, bid.get("comments", ""))
        c.fill = fill; c.border = border; c.alignment = Alignment(wrap_text=True)
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    proj_name = pi.get("name", "estimate")
    safe_name = "".join(c for c in proj_name if c.isalnum() or c in " _-").strip() or "estimate"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{safe_name}.xlsx",
    )


# -------------------------------
# PDF Download Route
# -------------------------------

@app.route("/download_summary")
@login_required
def download_summary():
    """Return the last generated order summary PDF."""
    global pdf_buffer
    address_filename = session.get("last_address_filename", "order_summary")
    if pdf_buffer is not None:
        pdf_buffer.seek(0)
        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{address_filename}.pdf",
        )
    pdf_path = session.get("pdf_path")
    if pdf_path and os.path.exists(pdf_path):
        return send_file(
            pdf_path,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{address_filename}.pdf",
        )

    flash("No PDF available for download.", "warning")
    return redirect(url_for("material_list"))
# -------------------------------
# Login and Logout Routes
# -------------------------------

@app.route("/login", methods=["GET", "POST"])
@limiter.limit("5 per minute;10 per hour", methods=["POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()

        user = get_user(email)
        if not user or not user["active"]:
            flash("Unauthorized email.", "danger")
            return redirect(url_for("login"))

        code = str(random.randint(100000, 999999))
        session["login_pending"] = {
            "email": email,
            "code": code,
            "ts": time.time(),
        }

        msg = Message("Your Login Code", recipients=[email])
        msg.body = (
            f"Your Zamora Inventory login code is: {code}\n\n"
            "This code expires in 10 minutes. Do not share it."
        )
        try:
            mail.send(msg)
        except Exception as e:
            app.logger.error(f"Error sending login code: {e}")
            flash("Error sending email. Please try again later.", "danger")
            return redirect(url_for("login"))

        flash("A 6-digit login code has been sent to your email.", "info")
        return redirect(url_for("verify_code"))

    return render_app("login", {"loginUrl": url_for("login")}, nav_links=[])


@app.route("/verify_code", methods=["GET", "POST"])
def verify_code():
    pending = session.get("login_pending")
    if not pending:
        flash("No login in progress.", "warning")
        return redirect(url_for("login"))

    if request.method == "POST":
        entered = request.form.get("code", "").strip()
        email = pending["email"]

        if time.time() - pending["ts"] > 600:
            session.pop("login_pending", None)
            flash("Code expired, please request a new one.", "danger")
            return redirect(url_for("login"))

        if entered != pending["code"]:
            new_count = increment_failed_attempts(email)
            if new_count >= 5:
                set_user_active(email, 0)
                session.pop("login_pending", None)
                flash(
                    "Too many failed attempts. Your account has been deactivated. "
                    "Please contact an admin.",
                    "danger",
                )
                return redirect(url_for("login"))
            remaining = 5 - new_count
            flash(f"Invalid code. {remaining} attempt(s) remaining.", "danger")
            return redirect(url_for("verify_code"))

        # Success — clear pending code and open session
        session.pop("login_pending", None)
        reset_failed_attempts(email)
        log_login(email, request.remote_addr)

        user = get_user(email)
        session["email"] = email
        session["role"] = user["role"]
        session.permanent = True
        session["last_activity"] = time.time()

        flash("Login successful!", "success")
        return redirect(url_for("index"))

    return render_template("verify_code.html", email=pending["email"])


@app.route("/verify_login/<token>")
def verify_login(token):
    """Legacy magic-link route kept for backwards compatibility."""
    try:
        email = serializer.loads(token, salt="email-confirmation", max_age=600)
        user = get_user(email)
        if not user or not user["active"]:
            flash("This account is not authorised.", "danger")
            return redirect(url_for("login"))
        log_login(email, request.remote_addr)
        session["email"] = email
        session["role"] = user["role"]
        session.permanent = True
        session["last_activity"] = time.time()
        flash("Login successful!", "success")
        return redirect(url_for("index"))
    except Exception:
        flash("Invalid or expired login link.", "danger")
        return redirect(url_for("login"))


@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully.", "info")
    return redirect(url_for("login"))


# -------------------------------
# Admin Routes
# -------------------------------

@app.route("/admin/users")
@admin_required
def admin_users():
    users = list_users()
    return render_template(
        "admin_users.html",
        users=users,
        add_url=url_for("admin_add_user"),
        toggle_url=url_for("admin_toggle_user"),
        role_url=url_for("admin_set_role"),
    )


@app.route("/admin/users/add", methods=["POST"])
@admin_required
def admin_add_user():
    email = request.form.get("email", "").strip().lower()
    role = request.form.get("role", "user")
    if not email:
        flash("Email is required.", "danger")
    elif role not in ("user", "admin"):
        flash("Invalid role.", "danger")
    elif not add_user(email, role):
        flash(f"{email} is already in the whitelist.", "warning")
    else:
        flash(f"{email} added as {role}.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/toggle", methods=["POST"])
@admin_required
def admin_toggle_user():
    email = request.form.get("email", "").strip().lower()
    action = request.form.get("action", "")
    if not email or action not in ("activate", "deactivate"):
        flash("Invalid request.", "danger")
        return redirect(url_for("admin_users"))
    set_user_active(email, 1 if action == "activate" else 0)
    if action == "activate":
        reset_failed_attempts(email)
    flash(f"{email} {action}d.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/role", methods=["POST"])
@admin_required
def admin_set_role():
    email = request.form.get("email", "").strip().lower()
    role = request.form.get("role", "")
    if not email or role not in ("user", "admin"):
        flash("Invalid request.", "danger")
        return redirect(url_for("admin_users"))
    set_user_role(email, role)
    flash(f"{email} role set to {role}.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/migrate_templates", methods=["POST"])
@admin_required
def admin_migrate_templates():
    """One-time import of local JSON template files into Turso."""
    templates_dir = config.TEMPLATE_DATA_DIR
    migrated = skipped = errors = 0
    if os.path.isdir(templates_dir):
        for root, _dirs, files in os.walk(templates_dir):
            for fname in files:
                if not fname.endswith(".json"):
                    continue
                path = os.path.join(root, fname)
                rel_dir = os.path.relpath(root, templates_dir)
                folder = "" if rel_dir == "." else rel_dir
                tname  = os.path.splitext(fname)[0]
                try:
                    with open(path) as fh:
                        data_json = fh.read()
                    existing = get_template_db(tname, folder, session["email"], "admin")
                    if existing:
                        skipped += 1
                        continue
                    tid = save_template_db(tname, folder, session["email"], "admin", data_json)
                    if tid:
                        migrated += 1
                    else:
                        errors += 1
                except Exception as e:
                    app.logger.error(f"Migration error for {fname}: {e}")
                    errors += 1
    flash(
        f"Migration complete: {migrated} imported, {skipped} already existed, {errors} error(s).",
        "success" if errors == 0 else "warning",
    )
    return redirect(url_for("admin_users"))


@app.route("/admin/login_history")
@admin_required
def admin_login_history():
    email = request.args.get("email", "").strip().lower() or None
    history = get_login_history(email=email, limit=200)
    users = list_users()
    return render_template(
        "admin_login_history.html",
        history=history,
        filter_email=email or "",
        users=users,
    )

# -------------------------------
# PDF Upload Routes
# -------------------------------

@app.route("/upload_pdf", methods=["GET", "POST"])
@login_required
def upload_pdf():
    """
    GET  — show the upload page (history of imported docs + upload form)
    POST — accept a PDF, parse it, save to DB, return result
    """
    _SUPPLIERS = [
        {"code": "LPS",  "label": "Lion Plumbing Supply"},
        {"code": "BPS",  "label": "Berger Plumbing Supply"},
        {"code": "S2",   "label": "Supply 2"},
        {"code": "BOND", "label": "Bond Plumbing Supply"},
    ]

    if request.method == "GET":
        invoices = list_invoices()
        initial = {
            "invoices": invoices,
            "uploadUrl": url_for("upload_pdf"),
            "deleteUrl": url_for("delete_invoice_route"),
            "confirmUrl": url_for("confirm_upload"),
            "suppliers": _SUPPLIERS,
        }
        return render_app("upload_pdf", initial)

    # ── POST: parse only — does NOT save to DB ────────────────────────────
    if "pdf_file" not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400

    pdf_file = request.files["pdf_file"]
    if not pdf_file.filename or not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"success": False, "error": "Please upload a PDF file"}), 400

    supplier = request.form.get("supplier", "LPS")

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        pdf_file.save(tmp.name)
        tmp_path = tmp.name

    try:
        parsed = parse_pdf(tmp_path, supplier=supplier)
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        app.logger.error(f"PDF parse error: {e}")
        return jsonify({"success": False, "error": "Failed to parse PDF. Check server logs."}), 500
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    return jsonify({
        "success": True,
        "doc_type": parsed["doc_type"],
        "order_number": parsed["order_number"],
        "date": parsed["date"],
        "job_name": parsed["job_name"],
        "supplier": parsed["supplier"],
        "item_count": len(parsed["items"]),
        "items": parsed["items"],
    })


@app.route("/delete_invoice", methods=["POST"])
@login_required
def delete_invoice_route():
    invoice_id = request.form.get("invoice_id", type=int)
    if not invoice_id:
        flash("Invalid invoice ID.", "danger")
        return redirect(url_for("upload_pdf"))
    delete_invoice(invoice_id)
    flash("Document deleted.", "success")
    return redirect(url_for("upload_pdf"))


@app.route("/confirm_upload", methods=["POST"])
@login_required
def confirm_upload():
    """
    Accepts the user-reviewed parsed data as JSON and saves it to the DB.
    Body: { "parsed": {...doc fields + items...}, "filename": "..." }
    """
    body = request.get_json(force=True, silent=True) or {}
    app.logger.info(f"confirm_upload received: {body}")
    print(f"confirm_upload received: {body}")
    parsed = body.get("parsed")
    filename = body.get("filename", "")

    if not parsed or not isinstance(parsed, dict):
        return jsonify({"success": False, "error": "No parsed data provided."}), 400

    required = {"doc_type", "order_number", "items"}
    missing = required - parsed.keys()
    if missing:
        return jsonify({"success": False, "error": f"Missing fields: {missing}"}), 400

    invoice_id = save_parsed_document(parsed, filename=filename)

    if invoice_id == -1:
        return jsonify({
            "success": False,
            "duplicate": True,
            "error": f"Document {parsed['order_number']} has already been imported.",
        }), 409

    refresh_catalog()

    return jsonify({
        "success": True,
        "invoice_id": invoice_id,
        "doc_type": parsed["doc_type"],
        "order_number": parsed["order_number"],
        "date": parsed.get("date", ""),
        "job_name": parsed.get("job_name", ""),
        "item_count": len(parsed["items"]),
    })


# -------------------------------
# Run the Application
# -------------------------------

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000, debug=True)
